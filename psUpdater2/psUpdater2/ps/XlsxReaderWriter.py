import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from typing import Tuple, List, Optional, Dict, Union, Set, Any, Iterable

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ps.spec import ProgramSpec, Deployment, Playlist, Message, Recipient, General

ARTIFACTS = ['general', 'deployments', 'content', 'recipients']


class _XlsxExporter:
    def __init__(self, program_spec: ProgramSpec):
        self.program_id = program_spec.program_id
        self.program_spec = program_spec

    @staticmethod
    def pxl_columns_for(desired_columns: List[str], *, columns: Tuple[str]) -> List[int]:
        """
        Given some column names from a table, what will be the column numbers in a spreadsheet? 1-based.
        :param desired_columns: Columns for which the indices are desired.
        :param columns: All columns.
        :return: A list of the 1-based indices of the columns.
        """
        result: Set[int] = set()
        for ix, name in enumerate(columns):
            if name in desired_columns:
                result.add(ix + 1)
        return sorted(list(result))

    @staticmethod
    def _auto_width(ws: Worksheet, *, formats: Dict[int, Dict[str, Any]] = None) -> None:
        """
        Sets formatting for the worksheet. In the most basic usage, sets the widths based on data.
        :param ws: The worksheet to be formatted.
        :param formats: A Dict of formats { column : {format_spec: value, ...}, ...}
                If no formats are supplied, this will set the widths on data columns so that
                everything is visible.
        :return: None
        """

        def format_cell(c) -> None:
            """
            Sets the formatting for a single cell. Supports widths, text wrapping, and date formatting.
            :param c: The cell to be formattted.
            :return: nothing
            """
            if formats and c.column in formats:
                f = formats[c.column]
                if f.get('no_width'):
                    # don't say anything about the width
                    pass
                elif f.get('width') is not None:
                    # explicitly specified width
                    dims[c.cell.column_letter] = f.get('width')
                else:
                    # width based on content, with optional min and/or max width
                    dim = max((dims.get(c.column_letter, 0), len(str(c.value))))
                    if f.get('max_width') is not None:
                        dim = min(dim, f.get('max_width'))  # no more than this
                    if f.get('min_width') is not None:
                        dim = max(dim, f.get('min_width'))  # no less than this
                    dims[c.column_letter] = dim
                if f.get('wrap'):
                    c.alignment = Alignment(wrapText=True)
                if f.get('date'):
                    c.number_format = 'yyyy-mm-dd'
            else:
                dims[c.column_letter] = max((dims.get(c.column_letter, 0), len(str(c.value))))

        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    format_cell(cell)
        # Set the widths that were accumulated by examining the cell contents.
        for col, value in dims.items():
            ws.column_dimensions[col].width = value
        # Make the header row bold.
        bold: Font = Font(bold=True)
        for cell in ws[1:1]:
            cell.font = bold

    @staticmethod
    def _add_sheet(wb: Workbook, sheet_name: str, values: List[Tuple[str]],
                   header: Tuple[str]) -> Worksheet:
        """
        Adds one sheet to the workbook, for 'content', 'recipients', etc.
        :param wb: The workbook into which to add the sheet.
        :param sheet_name: Name to be given to the new sheet.
        :param values: The contents of the sheet.
        :param header: Values (names) for the column headers.
        :return: The new worksheet.
        """
        ws = wb.create_sheet(sheet_name)
        ws.append(list(header))
        for tup in values:
            ws.append(tup)
        return ws

    def _save_general(self, wb: Workbook) -> Worksheet:
        # If the program hasn't yet been fully initialized, there may not be the data for a "general" sheet.
        if self.program_spec.general is not None:
            columns = General.csv_header
            values = [self.program_spec.general.csv_row]
            ws = self._add_sheet(wb, sheet_name='General', values=values,
                                 header=columns)
            self._auto_width(ws)
            return ws

    def _save_deployments(self, wb: Workbook) -> Worksheet:
        """
        Saves the Deployments sheet, with formatting on the date columns.
        :param wb: The workbook to receive the sheet.
        :return: The new worksheet.
        """
        columns = Deployment.csv_header
        values = [d.csv_row for d in self.program_spec.deployments]
        ws = self._add_sheet(wb, sheet_name='Deployments', values=values, header=columns)
        formats = {c: {'date': True} for c in self.pxl_columns_for(['Start Date', 'End Date'], columns=columns)}
        self._auto_width(ws, formats=formats)
        return ws

    def _save_content(self, wb: Workbook) -> Worksheet:
        """
        Saves the Content sheet, with text wrapping on the title and key points.
        :param wb: The workbook to receive the sheet.
        :return: The new worksheet.
        """
        columns = Message.csv_header
        values = [msg.csv_row for depl in self.program_spec.deployments for pl in depl.playlists for msg in
                  pl.messages]
        ws = self._add_sheet(wb, sheet_name='Content', values=values, header=columns)
        formats = {c: {'max_width': 60, 'wrap': True} for c in
                   self.pxl_columns_for(["Message Title", "Key Points"], columns=columns)}
        self._auto_width(ws, formats=formats)
        return ws

    def _save_recipients(self, wb: Workbook) -> Worksheet:
        """
        Saves the Recipients sheet.
        :param wb: The workbook to receive the sheet.
        :return: The new worksheet.
        """
        columns = Recipient.csv_header
        values = [r.csv_row for r in self.program_spec.recipients]
        ws = self._add_sheet(wb, sheet_name='Recipients', values=values, header=columns)
        self._auto_width(ws)
        return ws

    @staticmethod
    def _get_csv(rows: Iterable[Tuple[str]], header: Tuple[str]) -> str:
        csv_io = StringIO()
        writer = csv.writer(csv_io)
        writer.writerow(header)
        for tup in rows:
            writer.writerow(tup)
        csv_str = csv_io.getvalue()
        return csv_str

    def get_spreadsheet(self) -> bytes:
        """
        Gets the raw spreadsheet data. Can be saved to a file, to an S3 bucket, or opened as a spreadsheet.
        :return: The bytes of the spreadsheet
        """
        # create the spreadsheet data
        wb: Workbook = Workbook()
        for s in wb.get_sheet_names():
            del wb[s]
        self._save_general(wb)
        self._save_deployments(wb)
        self._save_content(wb)
        self._save_recipients(wb)
        xls_data = BytesIO()
        wb.save(xls_data)
        data = xls_data.getvalue()
        return data

    def get_csv(self, artifact: str) -> Optional[str]:
        """
        Gets one of the spreadsheet tabs, as a .csv, in a string format.
        :param artifact: The desired tab.
        :return: The tab, as a .csv, as a string.
        """
        if artifact == 'general' and self.program_spec.general is not None:
            return self._get_csv([self.program_spec.general.csv_row], General.csv_header)
        elif artifact == 'deployments':
            return self._get_csv((x.csv_row for x in self.program_spec.deployments), Deployment.csv_header)
        elif artifact == 'content':
            return self._get_csv((x.csv_row for x in self.program_spec.content), Message.csv_header)
        elif artifact == 'recipients':
            return self._get_csv((x.csv_row for x in self.program_spec.recipients), Recipient.csv_header)
        else:
            return None


class _XlsxImporter:
    def __init__(self, programid: str, data_or_path: Union[bytes, Path]):
        self._programid: str = programid
        self._data_or_path: Union[bytes, Path] = data_or_path

        self._sheets: Optional[openpyxl.Workbook] = None

        self._deployments: List[Dict] = []
        self._content: List[Dict] = []
        self._recipients: List[Dict] = []
        self._general: Dict = {}

    def _load_sheets(self, data: bytes) -> Tuple[bool, List[str]]:
        result = []
        try:
            _bytes = BytesIO(data)
            self._sheets = openpyxl.load_workbook(_bytes)
            return True, []
        except Exception as ex:
            result.append(f'Exception loading workbook: {ex}')
        return False, result

    def _parse_sheets(self) -> Tuple[bool, List[str]]:
        def parse_sheet(sheet: Worksheet, dc:dataclass, additional_map: Dict[str, str] = None) -> \
                List[Dict[str, str]]:
            column_import_map = dc.csv_to_internal_map
            if additional_map is not None:
                column_import_map.update(additional_map)

            # Raw rows
            rows = [r for r in sheet]
            # header values from spreadsheet
            header = [r.value for r in rows[0]]
            rows = rows[1:]
            # what we'll call the values in the output dict. Unknown columns transfered as-is
            column_names = [column_import_map[h] if h in column_import_map else h for h in header]
            result = []
            for row in rows:
                # Excel will give us a row of empty values if we've looked at it hard. Skip rows with no data.
                # We don't do anything with formulae, and customers like to put a sum of talking books so skip those
                # as well.
                if all([cell.data_type in 'nf' for cell in row]):
                    continue

                # normalize
                values = [x.value.strip() if isinstance(x.value, str) else x.value for x in row]
                result.append(dict(zip(column_names, values)))
            return result

        try:
            additional_message_mappings = {'Deployment #': 'deployment_num', 'Playlist Title': 'playlist_title'}
            additional_recipient_mappings = {'Model': 'listening_model', 'Language': 'language'}
            self._deployments = parse_sheet(self._sheets['Deployments'], Deployment)
            self._content = parse_sheet(self._sheets['Content'], Message,
                                        additional_map=additional_message_mappings)
            if 'Recipients' in self._sheets:
                self._recipients = parse_sheet(self._sheets['Recipients'], Recipient,
                                               additional_map=additional_recipient_mappings)
            elif 'Components' in self._sheets:
                names_dict = parse_sheet(self._sheets['Components'], {'component': 'Component'})
                names = [x['component'] for x in names_dict]
                self._recipients = []
                for name in names:
                    self._recipients.extend(parse_sheet(self._sheets[name], Recipient,
                                                        additional_map=additional_recipient_mappings))

            if 'General' in self._sheets:
                self._general = parse_sheet(self._sheets['General'], General)[0]

            return True, []
        except Exception as ex:
            return False, [f'Exception parsing workbook: {ex}']

    def _add_sheets(self) -> Tuple[ProgramSpec, List[str]]:
        ok = True
        messages: List[str] = []
        deployments: Dict[int, Deployment] = {}
        playlists: Dict[int, Dict[str, Playlist]] = {}
        prog_spec: ProgramSpec = ProgramSpec(self._programid)

        for d in self._deployments:
            depl = prog_spec.add_deployment(d)
            deployments[depl.deploymentnumber] = depl
            playlists[depl.deploymentnumber] = {}

        message_export_map = Message.internal_to_csv_map
        for c in self._content:
            msg_title = c['title']
            depl_num = c['deployment_num']
            pl_title = c['playlist_title']
            if depl_num not in deployments:
                ok = False
                messages.append(
                    f'    Message "{msg_title}" requires {message_export_map["deployment_num"]} {depl_num}' +
                    f', which is not in the Deployments sheet.')
            else:
                depl = deployments[depl_num]
                pl_in_depl = playlists[depl_num]
                pl = pl_in_depl.get(pl_title)
                if pl is None:
                    pl = depl.add_playlist({'title': pl_title})
                    pl_in_depl[pl_title] = pl
                pl.add_message(c)

        for r in self._recipients:
            r['program_id'] = self._programid
            prog_spec.add_recipient(r)

        # Older speradsheets had a different 'General' tab, with none of the current general info. Ignore those, if found.
        if self._general and all([x in self._general for x in General.required_columns]):
            prog_spec.add_general(self._general)

        return prog_spec if ok else None, messages

    def do_import(self) -> Tuple[Union[bool, Optional[ProgramSpec]], List[str]]:
        if isinstance(self._data_or_path, Path):
            with self._data_or_path.open('rb') as input_file:
                data = input_file.read()
        else:
            data = self._data_or_path
        messages: List[str]
        result, messages = self._load_sheets(data)
        if result:
            result, msgs = self._parse_sheets()
            messages.extend(msgs)
            if result:
                result, msgs = self._add_sheets()
                messages.extend(msgs)

        return result, messages


def write_to_xlsx(program_spec: ProgramSpec, path: Optional[Path] = None) -> bytes:
    """
    Exports a Program Specification to a spreadsheet.

    :param program_spec: The Program Spec to be exported.
    :param path: An optional Path to which the spreadsheet file will be written.
    :return: The bytes of the spreadsheet.
    """
    data: bytes = _XlsxExporter(program_spec).get_spreadsheet()
    # Write to a file if desired
    if path is not None:
        if path.is_dir():
            path = path_for_csv(path, 'progspec')
        with path.open(mode='wb') as xls_file:
            xls_file.write(data)
    return data


def path_for_csv(out_dir: Path, artifact: str, prefix: str = '') -> Path:
    if artifact == 'progspec':
        return Path(out_dir, f'{prefix}{artifact}.xlsx')
    return Path(out_dir, f'{prefix}{artifact}.csv')


def write_to_csv(program_spec: ProgramSpec, artifact: str, path: Optional[Path] = None, prefix: str = '') -> Optional[
    str]:
    """
    Export an artifact from a Program Specification. The artifact is one of ['general', 'deployments', 'content', 'recipients'],
    the three sheets of the Program Specification spreadsheet, and the three csv files used by the ACM.

    :param program_spec: The Program Specification from which a csv is to be exported.
    :param artifact: The artifact to be exported.
    :param path: An optional Path to which the csv will be written.
    :return: The csv data, as a string.
    """
    data: Optional[str] = _XlsxExporter(program_spec).get_csv(artifact)
    if path is not None and data is not None:
        if path.is_dir():
            path = path_for_csv(path, artifact, prefix)
        with path.open(mode='w') as csv_file:
            csv_file.write(data)
    return data


def read_from_xlsx(programid: str, data_or_path: Union[bytes, Path]) -> \
        Tuple[Union[bool, Optional[ProgramSpec]], List[str]]:
    """
    Imports a program specification from a spreadsheet.

    :param programid: The programid of the imported program specification.
    :param data_or_path: Either the bytes of a spreadsheet, or a Path to a spreadsheet file.
    :return: a Tuple consisting of:
        either False or a Program Specification
        A list of any messages generated by the import operation.
    """
    return _XlsxImporter(programid, data_or_path).do_import()
