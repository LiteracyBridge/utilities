import csv
import dataclasses
from io import BytesIO, StringIO
from typing import List, Set, Dict, Any, Optional, Tuple, Union

from openpyxl.styles import Alignment, Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

import Spec


class ExportProcessor:
    def __init__(self, program_spec: Spec.Program):
        self.program_id = program_spec.program_id
        self.program_spec = program_spec

    @staticmethod
    def pxl_columns_for(desired_columns: List[str], *, columns: List[str]) -> List[int]:
        """
        Given some column names from a table, what will be the column numbers in a spreadsheet? 1-based.
        :param desired_columns: Columns for which the indices are desired.
        :param columns: All columns.
        :return: A list of the 1-based indices of the columns.
        """
        result: Set[int] = set()
        for ix, name in enumerate(columns):
            if name in desired_columns:
                result.add(ix + 1)
        return sorted(list(result))

    @staticmethod
    def _auto_width(ws: Worksheet, *, formats: Dict[int, Dict[str, Any]] = None) -> None:
        """
        Sets formatting for the worksheet. In the most basic usage, sets the widths based on data.
        :param ws: The worksheet to be formatted.
        :param formats: A Dict of formats { column : {format_spec: value, ...}, ...}
                If no formats are supplied, this will set the widths on data columns so that
                everything is visible.
        :return: None
        """

        def format_cell(cell) -> None:
            """
            Sets the formatting for a single cell. Supports widths, text wrapping, and date formatting.
            :param cell: The cell to be formattted.
            :return: nothing
            """
            if formats and cell.column in formats:
                format = formats[cell.column]
                if format.get('no_width'):
                    # don't say anything about the width
                    pass
                elif format.get('width') is not None:
                    # explicitly specified width
                    dims[cell.cell.column_letter] = format.get('width')
                else:
                    # width based on content, with optional min and/or max width
                    dim = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                    if format.get('max_width') is not None:
                        dim = min(dim, format.get('max_width'))  # no more than this
                    if format.get('min_width') is not None:
                        dim = max(dim, format.get('min_width'))  # no less than this
                    dims[cell.column_letter] = dim
                if format.get('wrap'):
                    cell.alignment = Alignment(wrapText=True)
                if format.get('date'):
                    cell.number_format = 'yyyy-mm-dd'
            else:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))

        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    format_cell(cell)
        # Set the widths that were accumulated by examining the cell contents.
        for col, value in dims.items():
            ws.column_dimensions[col].width = value
        # Make the header row bold.
        bold: Font = Font(bold=True)
        for cell in ws[1:1]:
            cell.font = bold

    @staticmethod
    def _add_sheet(wb: Workbook, sheet_name: str, values: List[Union[dict, dataclasses.dataclass]],
                   header: List[str]) -> Worksheet:
        """
        Adds one sheet to the workbook, for 'content', 'recipients', etc.
        :param wb: The workbook into which to add the sheet.
        :param sheet_name: Name to be given to the new sheet.
        :param values: The contents of the sheet.
        :param header: Values (names) for the column headers.
        :return: The new worksheet.
        """
        ws = wb.create_sheet(sheet_name)
        ws.append(list(header))
        for row in values:
            if dataclasses.is_dataclass(row):
                tup = dataclasses.astuple(row)
            else:
                tup = tuple(row.values())
            ws.append(tup)
        return ws

    def _save_general(self, wb: Workbook) -> Worksheet:
        # If the program hasn't yet been fully initialized, there may not be the data for a "general" sheet.
        if self.program_spec._program is not None:
            ws = self._add_sheet(wb, sheet_name='General', values=[self.program_spec.program],
                                 header=Spec.general_sql_2_csv.values())
            self._auto_width(ws)
            return ws

    def _get_deployments(self, for_csv:bool = False) -> Tuple[List[str], List[Dict[str, str]]]:
        if for_csv:
            columns = list(Spec.deployment_sql_2_csv.keys()) + ['deployment']
            values = [{k: v for k, v in dataclasses.asdict(x).items() if k in columns} for x in
                  self.program_spec.deployments]
        else:
            columns = Spec.deployment_sql_2_csv.values()
            values = [{k: v for k, v in dataclasses.asdict(x).items() if k in Spec.deployment_sql_2_csv} for x in
                  self.program_spec.deployments]
        return columns, values

    def _save_deployments(self, wb: Workbook) -> Worksheet:
        """
        Saves the Deployments sheet, with formatting on the date columns.
        :param wb: The workbook to receive the sheet.
        :return: The new worksheet.
        """
        columns, values = self._get_deployments()
        ws = self._add_sheet(wb, sheet_name='Deployments', values=values, header=columns)
        formats = {c: {'date': True} for c in self.pxl_columns_for(['Start Date', 'End Date'], columns=columns)}
        self._auto_width(ws, formats=formats)
        return ws

    def _save_content(self, wb: Workbook) -> Worksheet:
        """
        Saves the Content sheet, with text wrapping on the title and key points.
        :param wb: The workbook to receive the sheet.
        :return: The new worksheet.
        """
        columns = Spec.content_sql_2_csv.values()
        ws = self._add_sheet(wb, sheet_name='Content', values=self.program_spec.content, header=columns)
        formats = {c: {'max_width': 60, 'wrap': True} for c in
                   self.pxl_columns_for(["Message Title", "Key Points"], columns=columns)}
        self._auto_width(ws, formats=formats)
        return ws

    def _save_recipients(self, wb: Workbook) -> Worksheet:
        """
        Saves the Recipients sheet.
        :param wb: The workbook to receive the sheet.
        :return: The new worksheet.
        """
        columns = Spec.recipient_sql_2_csv.values()
        ws = self._add_sheet(wb, sheet_name='Recipients', values=self.program_spec.recipients, header=columns)
        self._auto_width(ws)
        return ws

    def _get_csv(self, rows: List[Union[dict, dataclasses.dataclass]], name: str, header: List[str]) -> str:
        csv_io = StringIO()
        writer = csv.writer(csv_io)
        writer.writerow(header)
        for row in rows:
            if dataclasses.is_dataclass(row):
                tup = dataclasses.astuple(row)
            else:
                tup = tuple(row.values())
            writer.writerow(tup)
        csv_str = csv_io.getvalue()
        return csv_str

    def get_spreadsheet(self) -> bytes:
        """
        Gets the raw spreadsheet data. Can be saved to a file, to an S3 bucket, or opened as a spreadsheet.
        :return: The bytes of the spreadsheet
        """
        # create the spreadsheet data
        wb: Workbook = Workbook()
        for s in wb.get_sheet_names():
            del wb[s]
        self._save_general(wb)
        self._save_deployments(wb)
        self._save_content(wb)
        self._save_recipients(wb)
        xls_data = BytesIO()
        wb.save(xls_data)
        data = xls_data.getvalue()
        return data

    def get_csv(self, artifact: str) -> Optional[str]:
        """
        Gets one of the spreadsheet tabs, as a .csv, in a string format.
        :param artifact: The desired tab.
        :return: The tab, as a .csv, as a string.
        """
        if artifact == 'general' and self.program_spec.program is not None:
            return self._get_csv([self.program_spec.program], 'general', Spec.general_sql_2_csv.keys())
        elif artifact == 'deployments':
            columns, values = self._get_deployments(for_csv=True)
            return self._get_csv(values, 'deployments', columns)
        elif artifact == 'content':
            return self._get_csv(self.program_spec.content, 'content', Spec.content_sql_2_csv.keys())
        elif artifact == 'recipients':
            return self._get_csv(self.program_spec.recipients, 'recipients', Spec.recipient_sql_2_csv.keys())
        else:
            return None
