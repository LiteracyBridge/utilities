# import pandas as pd
import sys
import traceback

# noinspection PyUnresolvedReferences
from openpyxl import load_workbook

from . import errors, utils
from .programspec_constants import GENERAL, CONTENT, DEPLOYMENTS, \
    COMPONENTS, required_sheets, \
    required_columns, required_data, optional_columns, columns_to_members_map, \
    string_columns, column_names_to_member_names, optional_sheets, default_data, columns_to_rename, columns_to_remove

'''
Interface to Program Specification spreadsheet.
'''

options = {}

# { '# TBs': 'num_tbs', ...}
column_name_to_property_name_map = columns_to_members_map(GENERAL, CONTENT, DEPLOYMENTS, COMPONENTS, 'recipient')
# { 'num_tbs': '# TBs', ...}
property_name_to_column_name_map = {v: k for k, v in column_name_to_property_name_map.items()}



# "Sparse" list used for spreadsheet rows, since "a few" rows on the sheet may be blank (or skipped),
# and yet it is useful to keep track of which row data came from, in order to update the sheet.
# Taken from this SO post: https://stackoverflow.com/questions/1857780/sparse-assignment-list-in-python
class SparseList(list):
    def __setitem__(self, index, value):
        missing = index - len(self) + 1
        if missing > 0:
            self.extend([None] * missing)
        list.__setitem__(self, index, value)

    def __getitem__(self, index):
        try:
            return list.__getitem__(self, index)
        except IndexError:
            return None


class Spreadsheet:
    def __init__(self, filename):
        self._wb = None
        self._has_changes = False
        self._changes = []

        # cached column positions, col_idx-1 and column
        # { sheet : {column : col_idx-1}, ...}
        self._indices = {}  # 0-based index, not 1-based as col_idx
        # cache for { sheet : [optional_member_name, ...]}
        self._optional_member_names = {}

        # General info, (partner, program, num_deployments)
        self._general_info = None
        # These two are inverses of each other; a two-way map.
        self._aliases = {}
        self._aliased = {}
        # Map of deployments as { depl# : (start, end, {filtered_property:filter}) }
        self._deployments = None
        # List of component names, [component, ...]
        self._components = None
        # Map of of { component_name : [Recipient, ...] }
        self._recipients = None
        # Recipient ids that we've seen. Its an error if there are duplicates.
        self._recipientids = set()
        # Array by deployment of array by playlist of content
        self._content = None

        self._filename = filename
        self._wb = load_workbook(filename)

        self._validate()

    @property
    def ok(self):
        return not errors.has_error()

    @property
    def general_info(self):
        return self._general_info

    @property
    def deployments(self):
        return self._deployments

    @property
    def components(self):
        return self._components

    @property
    def recipients(self):
        return self._recipients

    @property
    def content(self):
        return self._content

    @property
    def recipientids(self):
        return self._recipientids

    # Save the data to the given file.
    def save(self, filename):
        self._wb.save(filename)

    # Store a value into the spreadsheet. This is tailored to the Program Specification,
    # and assumes that there is exactly one column for property_name.
    def store_value(self, sheet_name, row_number, property_name, value):
        sh = self._wb[sheet_name]
        column_name = property_name_to_column_name_map[property_name]
        indices = self._indices[sheet_name]

        # Add the column?
        if column_name not in indices:
            indices[column_name] = sh.max_column
            sh.cell(row=1, column=sh.max_column + 1, value=column_name)
            self._changes.append(('column', sh.max_column + 1, column_name))
        # Create or modify the cell.
        sh.cell(row=row_number, column=indices[column_name] + 1, value=value)
        # Log the change.
        self._changes.append(('cell', row_number, indices[column_name] + 1, value))

    def _sheet_type_for(self, sheet_name: str):
        '''
        Given a sheet name, return the sheet type. For sheets like 'General', or 'Content', the sheet type is
        also the sheet name, but for component recipient sheets, the name will be component's name. If the name
        matches a known component, then it is a 'recipient' type of sheet. Otherwise, we just hope that it
        is a recipient sheet (if it isn't it'll surely have lots of errors, so it is a very low probability of
        causing real problems.
        :param sheet_name: The name of the sheet.
        :return: The type of the sheet.
        '''
        if sheet_name in required_columns:
            return sheet_name
        if sheet_name in self._components:
            return 'recipient'
        # We don't know what it is; hope it is recipient?
        return 'recipient'

    # Given a row from a sheet of type sheet_type, extract any optional data into a dictionary.
    def _extract_optional_members(self, row_dict, sheet_type: str):
        # lazy init { sheet_type : [ optional_prop1, optional_prop2, ... ] }
        if sheet_type not in self._optional_member_names:
            self._optional_member_names[sheet_type] = [column_name_to_property_name_map[c] for c in
                                                       optional_columns[sheet_type]]
        optional_member_names = self._optional_member_names[sheet_type]
        optionals = {k: row_dict[k] for k in optional_member_names if k in row_dict and row_dict[k] is not None}
        return optionals

    # Reads a sheet named 'sheet_name' from the workbook. The type of the sheet may be passed in sheet_type; if
    # no sheet_type is provided, sheet_name is used for the type.
    #
    # Required columns are given by the global required_columns[sheet_type], and optional ones
    # by optional_columns[sheet_type].
    #
    # The result will be a list of:
    # - a dictionary of {prop_name: value}, plus {'row_num': row_num} of the excel row number.
    # - empty rows are ignored
    # - missing required data generates an error. Rows with missing data are skipped unless
    #   include_partial_rows is True
    #
    # Row numbers start with 1.
    def _get_rows_for_sheet(self, sheet_name: str, sheet_type: str = None, include_partial_rows: bool = False):
        sheet_type = sheet_type or sheet_name # a known name, or of type 'recipient'

        # If there is no such sheet, but there is default data for the sheet, return that.
        if sheet_type in default_data and sheet_name not in self._wb.sheetnames:
            result = default_data[sheet_name]
            return result

        # which columns and where they are in this sheet, as {column_name: column_ix}
        indices = self._indices[sheet_name]
        # [(prop_name, column_ix)] of the columns we must have, the required columns.
        name_and_ix = [(column_name_to_property_name_map[c], indices[c]) for c in required_columns[sheet_type]]
        # Add the columns we may or may not have, the optional columns.
        if sheet_type in optional_columns:
            for c in optional_columns[sheet_type]:
                if c in indices:
                    name_and_ix.append((column_name_to_property_name_map[c], indices[c]))

        # 'required_data', if it exists, is a subset of the required columns. If no 'required_data', all required
        # columns must have data. Build a [(column_name, column_ix)] of the required data. Note column vs prop
        # name; that's because we report on missing data, and need the name by which user knows it.
        required_cells = required_data.get(sheet_type, None) or required_columns[sheet_type]
        required_cell_indices = [(self._indices[sheet_name][column], column) for column in required_cells]

        # Columns required to be strings.
        string_cells = column_names_to_member_names(string_columns[sheet_type]).values()
        def normalize(x, name):
            if x is not None and name in string_cells and not isinstance(x, str):
                x = str(x)
            if isinstance(x, str):
                return x.strip()
            return x

        # Get the spreadsheet data.
        sh = self._wb[sheet_name]
        result = []

        # Getting a slice of a single row returns the row, not a list with one element, but we always want a list. So,
        # get a slice of one more than we want, which will always be a list, then slice that down to what we do want.
        #
        # But, if there is only one row (the header, with no content), we can't get it in a list. So, detect that
        # and return the empty list.
        if sh.max_row == 1:
            return result
        
        raw_rows = sh[2:sh.max_row + 1]
        raw_rows = raw_rows[0:len(raw_rows) - 1]

        # Finally ready to gather the actual data.
        for ix in range(0, len(raw_rows)):
            row = raw_rows[ix]
            row_num = ix + 2
            # Excel will give us a row of empty values if we've looked at it hard. Skip rows with no data.
            if all([cell.value is None for cell in row]):
                continue
            # See if there are columns with missing data (NB, not missing columns, but empty cells); report them by name.
            missing = [reqd_cell[1] for reqd_cell in required_cell_indices if row[reqd_cell[0]].value is None]
            if len(missing) > 0:
                errors.error(errors.missing_sheet_values,
                          {'columns': '", "'.join(missing), 'sheet': sheet_name, 'row': row_num})
            if len(missing) == 0 or include_partial_rows:
                # We have what we need; add to the result.
                # Build list of [dictionary of {member_name: column.value}]
                row_dict = {name: (normalize(row[cell_ix].value, name) if cell_ix is not None else None) for name, cell_ix in
                            name_and_ix}
                row_dict['row_num'] = row_num
                result.append(row_dict)

        return result

    # Performs fixups on sheets -- renaming or removing columns
    def _fixup_sheet_columns(self, sheet_name, sheet_type=None):
        ws = self._wb[sheet_name]
        sheet_type = sheet_type or sheet_name

        # Are there columns to be deleted? Delete them now.
        if sheet_type in columns_to_remove:
            cells = [x for x in ws[1] if x.value in columns_to_remove[sheet_type]]
            if len(cells) > 0:
                for cell in cells:
                    ws.delete_cols(cell.col_idx)
                # for cell in sheet[1]:
                #     if cell.value:
                #         letter = cell.column_letter
                #         sheet.column_dimensions[letter].bestFit = True

        # Are there renames for this sheet? Do then now, so when we collect data, we have the proper names.
        if sheet_type in columns_to_rename:
            renames = columns_to_rename[sheet_type]
            # Rename any indicated columns
            for cell in ws[1]:
                if cell.value in renames:
                    cell.value = renames[cell.value]
                    self._changes.append(('heading', 1, cell.col_idx, cell.value))

    #
    # Validates that the given sheet has exactly one of every required column. Checks optional columns, and if any
    # are present, must be exactly one of them. Required columns are described by the global
    #   require_columns[sheet_type], and optional columns by the global optional_columns[sheet_type].
    # If sheet_type is not explicitly provided, sheet_name is used.
    #
    # Populates _indices[sheet_type] and _columns[sheet_type] with the col_idx and column of every required column and
    # any present optional columns.
    def _validate_sheet_columns(self, sheet_name, sheet_type=None):
        # Make any adjustments very early.
        self._fixup_sheet_columns(sheet_name, sheet_type)

        sheet = self._wb[sheet_name]
        sheet_type = sheet_type or sheet_name
        required = required_columns[sheet_type]
        columns_found = set()
        column_errors = set()
        indices = {}
        columns = {}
        self._indices[sheet.title] = indices

        # Get information about both required and optional columns. Later, check that we found all required columns.
        columns_of_interest = required
        if sheet_type in optional_columns:
            columns_of_interest = columns_of_interest + optional_columns[sheet_type]
        # for every cell in the top row...
        for cell in sheet[1]:
            name = cell.value
            # do we care about this column?
            if name in columns_of_interest:
                # It's an error to see any of "our" columns more than once. Check if we already have seen it.
                if name in columns_found:
                    if name in indices:
                        del indices[name]
                    if name in columns:
                        del columns[name]
                    if name not in column_errors:
                        errors.error(errors.duplicate_columns, {'column': name, 'sheet': sheet_name})
                        column_errors.add(name)
                else:
                    indices[name] = cell.col_idx - 1
                    columns[name] = cell.column
                    columns_found.add(name)
        args = {'sheet': sheet.title}
        self._no_missing(columns_found, required, errors.missing_columns_in_sheet, joiner='", "', args=args)

    # Validate that the workbook has the required sheets, and that the sheets have the required columns.
    def _validate_basic_sheets(self):
        # Don't need to check for duplicates; workbooks can't have 2 sheets with same name.
        wb_sheets = self._wb.sheetnames
        for sheet in required_sheets:
            if sheet not in wb_sheets:
                errors.error(errors.missing_sheet, {'sheet': sheet})
            else:
                self._validate_sheet_columns(sheet)
        for sheet in optional_sheets:
            if sheet in wb_sheets:
                self._validate_sheet_columns(sheet)

    # Currently no validation; just read it.
    def _validate_general_info(self):
        self._general_info = self._get_rows_for_sheet(GENERAL)[0]

    # Helper to check for dups, issue an error is any found.
    def _no_dups(self, iterable, message, joiner=', ', args=None):
        if args is None:
            args = {}
        # Get any items that appear more than once in iterable.
        dups = {item for item in iterable if iterable.count(item) > 1}
        if len(dups) > 0:
            args['duplicates'] = joiner.join([str(d) for d in dups])
            errors.error(message, args)

    # Helper to check that everything in an iterable is in a list of valid items
    def _all_valid(self, iterable, valid_list, message, joiner=', ', args=None):
        if args is None:
            args = {}
        # Get any items in iterable that are not in the valid_list.
        not_valid = [item for item in iterable if item not in valid_list]
        if len(not_valid) > 0:
            args['invalid'] = joiner.join([str(d) for d in not_valid])
            errors.error(message, args)

    # Helper to check that every item in required_list is present in iterable
    def _no_missing(self, iterable, required_list, message, joiner=', ', args=None):
        if args is None:
            args = {}
        # Get any items that are missing from the required list.
        missing = [reqd for reqd in required_list if reqd not in iterable]
        if len(missing) > 0:
            args['missing'] = joiner.join([str(m) for m in missing])
            errors.error(message, args)

    # Validate the Deployments sheet. As a side effect, caches the deployments and dates.
    def _validate_deployments_sheet(self):
        self._deployments = {}
        row_dicts = self._get_rows_for_sheet(DEPLOYMENTS)

        depl_numbers = [v['deployment_num'] for v in row_dicts]
        self._no_dups(depl_numbers, errors.duplicate_deployment_numbers)

        for row in row_dicts:
            filters = self._extract_optional_members(row, DEPLOYMENTS)
            deployment_row = utils.JsObject(row)
            if deployment_row.start_date >= deployment_row.end_date:
                errors.error(errors.deployment_start_after_end,
                         {'start': deployment_row.start_date, 'end': deployment_row.end_date,
                          'deployment': deployment_row.deployment_num})
            if deployment_row.deployment_num < 0:
                errors.error(errors.deployment_lt_zero, {'deployment': deployment_row.deployment_num})
            else:
                self._deployments[deployment_row.deployment_num] = (
                    deployment_row.start_date, deployment_row.end_date, filters)

        if len(depl_numbers) == 0:
            errors.error(errors.no_deployments)
        else:
            self._no_missing(self._deployments, range(min(depl_numbers), max(depl_numbers) + 1),
                         errors.missing_deployment_numbers)

    # Validates the contents of the Components sheet. Caches the component names.
    def _validate_components_sheet(self):
        row_dicts = self._get_rows_for_sheet(COMPONENTS)
        filtered_comps = [v['component'] for v in row_dicts]

        self._components = filtered_comps

        # There should be a sheet for every unique Component in the Components sheet
        wb_sheets = self._wb.sheetnames
        for name in filtered_comps:
            if name not in wb_sheets:
                errors.error(errors.missing_sheet, {'sheet': name})
            else:
                self._validate_sheet_columns(name, sheet_type='recipient')

    # Validate the content sheet, ie, the content calendar.
    def _validate_content_sheet(self):
        content = []
        row_dicts = self._get_rows_for_sheet(CONTENT)
        for row in row_dicts:
            filters = self._extract_optional_members(row, CONTENT)
            content_row = utils.JsObject(row)
            content_row.filters = filters
            depl = content_row.deployment_num
            if depl not in self._deployments:
                errors.error(errors.message_unknown_deployment,
                         {'row': content_row.row_num, 'deployment': depl,
                          'title': content_row.message_title})  # MESSAGE_TITLE
            content.append(content_row)

        self._content = content

    def _check_for_duplicate_recipients(self):
        def upr(x):
            return x.upper() if type(x) == str else x

        recipients_seen = {}
        for component_name in self._components:
            recipients = self._recipients[component_name]
            for recipient in recipients:
                recipientid = recipient['recipientid'] if 'recipientid' in recipient else None
                if recipientid:
                    if recipientid in self._recipientids:
                        errors.error(errors.duplicate_recipientids, {'recipientid': recipientid})
                    else:
                        self._recipientids.add(recipientid)
                key = (upr(recipient['community']), upr(recipient['group_name']), upr(recipient['agent']),
                       upr(recipient['language_code']))
                if key in recipients_seen:
                    repeat_of = recipients_seen[key]
                    args = {'community': recipient['community'], 'group': recipient['group_name'],
                            'component': component_name, 'row': recipient['row_num'],
                            'agent': recipient['agent'],
                            'component2': repeat_of['component'], 'row2': repeat_of['row_num']}
                    errors.error(errors.repeated_community_group, args)
                else:
                    recipients_seen[key] = recipient

    # for each member of the list of recipients, set the component, project, and affiliate
    def _add_component_to_recipients(self, component_name):
        recipients = self._recipients[component_name]
        for recipient in recipients:
            recipient['component'] = component_name;

    # For each of the components, validate the recipients
    def _validate_recipients(self):
        self._recipients = {}
        for component_name in self._components:
            if component_name in self._indices:
                self._recipients[component_name] = self._get_rows_for_sheet(component_name, sheet_type='recipient')
                self._add_component_to_recipients(component_name)
        self._check_for_duplicate_recipients()

    # performs validations on the workbook. Opportunistically caches data structures as it does so.
    def _validate(self):
        # noinspection PyBroadException
        try:
            components_ok = False
            self._validate_basic_sheets()  # loads components & their deployments
            self._validate_general_info()
            if self.ok:
                self._validate_deployments_sheet()
            if self.ok:
                self._validate_components_sheet()
                components_ok = self.ok
            self._validate_content_sheet()
            if components_ok:
                self._validate_recipients()
            if self.ok:
                self._fixup_sheet_columns('RECIP-TEMPLATE', 'recipient')
        except:
            exc_list = traceback.format_exception(*sys.exc_info())
            exc_list[-1] = exc_list[-1].rstrip()
            exc_list = [v.replace('\n', '\n  :  ') for v in exc_list]
            exc_msg = ''.join(exc_list)
            errors.error(errors.validation_exception, {'message': exc_msg})


def load(name):
    reader = Spreadsheet(name)
    return reader


if __name__ == "__main__":
    pass
